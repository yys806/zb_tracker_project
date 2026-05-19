from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_sheet(ws, max_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def main() -> None:
    out = Path("docs/wiring_table.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "接线表"

    headers = ["序号", "模块/器件", "端口/引脚", "连接到", "线材/颜色建议", "作用", "注意事项", "调试验证"]
    rows = [
        [1, "PCA9685", "VCC", "OrangePi 40Pin 第 1 脚 3.3V", "红色杜邦线", "给 PCA9685 I2C 逻辑部分供电", "这里是逻辑电源，不是舵机大电流电源", "I2C 扫描能看到 0x40"],
        [2, "PCA9685", "GND", "OrangePi 40Pin 第 6 脚 GND", "黑色杜邦线", "信号共地", "必须和 OrangePi、舵机电源共地，否则 PWM 信号不稳定", "I2C 扫描正常，舵机不乱抖"],
        [3, "PCA9685", "SDA", "OrangePi 40Pin 第 3 脚 SDA", "绿色/蓝色杜邦线", "I2C 数据线", "不要接到普通 GPIO；线不要太长", "scan_i2c.py 显示 0x40"],
        [4, "PCA9685", "SCL", "OrangePi 40Pin 第 5 脚 SCL", "黄色杜邦线", "I2C 时钟线", "SDA/SCL 不要反接", "scan_i2c.py 显示 0x40"],
        [5, "PCA9685", "V+ / USB / 绿色端子 5V+", "独立 5V 舵机电源正极", "红色电源线", "给舵机供电", "不建议从 OrangePi USB 口给舵机供电，电流可能不够", "舵机动作有力，不导致 OrangePi 重启"],
        [6, "PCA9685", "V+ / USB / 绿色端子 GND", "独立 5V 舵机电源负极，同时与 OrangePi GND 共地", "黑色电源线", "舵机电源回路和控制信号参考地", "一定要共地；只接 5V 不共地会导致舵机不受控", "舵机按命令转动，不乱跳"],
        [7, "俯仰舵机 tilt", "三线插头 信号/正极/负极", "PCA9685 channel 0", "信号线按板子标注接 PWM/S；红线接 V+；棕/黑接 GND", "控制摄像头上下俯仰", "确认插头方向，信号线接外侧/标 S 的一排", "servo-test 中 tilt=35/145 时上下运动"],
        [8, "水平舵机 pan", "三线插头 信号/正极/负极", "PCA9685 channel 1", "信号线按板子标注接 PWM/S；红线接 V+；棕/黑接 GND", "控制底座左右旋转", "当前配置 pan_channel=1", "servo-test 中 pan=20/160 时左右运动"],
        [9, "USB 摄像头", "USB-A 插头", "OrangePi USB 接口", "USB 线", "采集实时画面", "优先插稳定的 USB 口；摄像头松动会导致画面读取失败", "ls /dev/video* 能看到 /dev/video0"],
        [10, "OrangePi", "Type-C 电源口", "OrangePi 官方/稳定电源适配器", "Type-C 电源线", "给 OrangePi 主板供电", "不要把舵机大电流全部压在 OrangePi 上", "系统稳定，不因舵机动作重启"],
        [11, "OrangePi", "HDMI", "显示器 HDMI 输入", "HDMI 线", "本地显示和调试", "笔记本 HDMI 通常是输出口，不能当显示器输入", "显示器能看到 OrangePi 桌面/终端"],
        [12, "键盘/鼠标", "USB", "OrangePi USB 接口", "USB 线或无线接收器", "本地输入控制", "初次调试建议直接接键鼠和显示器", "能在 OrangePi 终端输入命令"],
    ]

    ws.append(headers)
    for row in rows:
        ws.append(row)

    style_sheet(ws, len(headers))

    widths = [8, 18, 24, 38, 24, 28, 42, 34]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 48
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    note = wb.create_sheet("安全检查")
    note_rows = [
        ["检查项", "通过标准"],
        ["PCA9685 逻辑电源", "VCC 接 3.3V，GND 与 OrangePi 共地"],
        ["舵机电源", "V+ 使用独立 5V 供电，不从 OrangePi USB 硬扛两个舵机"],
        ["I2C 扫描", "/dev/i2c-1 能看到 0x40 和 0x70"],
        ["俯仰舵机", "channel 0 对应上下运动，不顶机械限位"],
        ["水平舵机", "channel 1 对应左右运动，不顶机械限位"],
        ["摄像头", "ls /dev/video* 能看到视频设备，camera-test 有画面"],
    ]
    for row in note_rows:
        note.append(row)
    style_sheet(note, 2)
    note.column_dimensions["A"].width = 24
    note.column_dimensions["B"].width = 70
    note.freeze_panes = "A2"

    wb.save(out)

    check = load_workbook(out)
    print(out.resolve())
    print(check.sheetnames)
    print(check["接线表"].max_row, check["接线表"].max_column)


if __name__ == "__main__":
    main()
